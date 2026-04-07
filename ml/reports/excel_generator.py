"""
ml/reports/excel_generator.py — OpenPyXL Excel Report Generator
Sheet 1: Summary dashboard + bar chart
Sheet 2: Critical issues (red)
Sheet 3: Major issues (orange)
Sheet 4: All issues (colour-coded)
"""

import io
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                               PatternFill, Side)
from openpyxl.utils import get_column_letter

# ── Colour fills ───────────────────────────────────────────────────────────────
FILL_CRITICAL   = PatternFill("solid", fgColor="FADBD8")
FILL_MAJOR      = PatternFill("solid", fgColor="FAE5D3")
FILL_MINOR      = PatternFill("solid", fgColor="FEF9E7")
FILL_HEADER     = PatternFill("solid", fgColor="1B4F72")
FILL_SUBHEADER  = PatternFill("solid", fgColor="2E86C1")
FILL_PASS       = PatternFill("solid", fgColor="EAFAF1")
FILL_FAIL       = PatternFill("solid", fgColor="FDEDEC")
FILL_VARROC     = PatternFill("solid", fgColor="003366")
FILL_ORANGE     = PatternFill("solid", fgColor="E87722")
FILL_ALT        = PatternFill("solid", fgColor="EBF5FB")

FONT_WHITE_HDR  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
FONT_WHITE      = Font(name='Calibri', color="FFFFFF", size=10)
FONT_BOLD       = Font(name='Calibri', bold=True, size=10)
FONT_NORMAL     = Font(name='Calibri', size=10)
FONT_SMALL      = Font(name='Calibri', size=9)
FONT_CRITICAL   = Font(name='Calibri', bold=True, color="C0392B", size=10)
FONT_MAJOR      = Font(name='Calibri', bold=True, color="E67E22", size=10)
FONT_MINOR      = Font(name='Calibri', bold=True, color="D4AC0D", size=10)

THIN_BORDER = Border(
    left=Side(style='thin', color='BFC9CA'),
    right=Side(style='thin', color='BFC9CA'),
    top=Side(style='thin', color='BFC9CA'),
    bottom=Side(style='thin', color='BFC9CA'),
)
MEDIUM_BORDER = Border(
    left=Side(style='medium', color='1B4F72'),
    right=Side(style='medium', color='1B4F72'),
    top=Side(style='medium', color='1B4F72'),
    bottom=Side(style='medium', color='1B4F72'),
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='top', wrap_text=True)


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    return cell


def _col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ── Violation table headers & columns ─────────────────────────────────────────
VIOLATION_COLS = [
    ("Violation ID",         20),
    ("Entity Type",          12),
    ("Layer",                14),
    ("Rule ID",              16),
    ("Rule Name",            28),
    ("Severity",             12),
    ("Confidence",           12),
    ("ML Flagged",           12),
    ("Blocks Sign-Off",      14),
    ("Description",          45),
    ("LLM Suggestion",       45),
    ("Standard Citation",    22),
    ("Relevant Lesson ID",   18),
    ("Past Fix",             40),
    ("Risk Score",           12),
]


def _write_violation_headers(ws, row=1):
    for ci, (hdr, _) in enumerate(VIOLATION_COLS, start=1):
        _set_cell(ws, row, ci, hdr,
                  font=FONT_WHITE_HDR, fill=FILL_HEADER,
                  alignment=ALIGN_CENTER, border=THIN_BORDER)
    for ci, (_, w) in enumerate(VIOLATION_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    _row_height(ws, row, 22)


def _severity_fills(sev):
    return {
        'Critical': (FILL_CRITICAL, FONT_CRITICAL),
        'Major':    (FILL_MAJOR,    FONT_MAJOR),
        'Minor':    (FILL_MINOR,    FONT_MINOR),
    }.get(sev, (None, FONT_NORMAL))


def _write_violation_row(ws, row, v, alt=False):
    lesson = v.get('relevantLesson') or {}
    sev = v.get('severity', '')
    fill, sev_font = _severity_fills(sev)
    base_fill = fill if fill else (FILL_ALT if alt else None)

    values = [
        v.get('violationId', '—'),
        v.get('entity_type', '—'),
        v.get('layer', '—'),
        v.get('rule_id', '—'),
        v.get('ruleName', '—'),
        sev,
        v.get('confidenceScore', 0),
        'Yes' if v.get('mlFlagged') else 'No',
        'YES' if v.get('blocksSignOff') else 'No',
        v.get('violation_description', '—'),
        v.get('llmSuggestion') or '(pending)',
        v.get('standardCitation', '—'),
        lesson.get('violationId', '—'),
        lesson.get('how_it_was_fixed', '—'),
        lesson.get('riskScore', '—'),
    ]
    for ci, val in enumerate(values, start=1):
        font = sev_font if ci == 6 else FONT_NORMAL
        _set_cell(ws, row, ci, val,
                  font=font, fill=base_fill,
                  alignment=ALIGN_LEFT, border=THIN_BORDER)
    _row_height(ws, row, 40)


# ── Sheet 1: Summary Dashboard ────────────────────────────────────────────────
def build_summary_sheet(ws, data):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    counts  = data.get('counts', {})
    total_c = counts.get('critical', 0)
    total_m = counts.get('major', 0)
    total_n = counts.get('minorSuppressed', 0)
    blocks  = data.get('blocksSignOff', False)

    # Title banner
    ws.merge_cells('A1:H1')
    _set_cell(ws, 1, 1, "VARROC AI VALIDATION SYSTEM — DRAWING QUALITY REPORT",
              font=Font(name='Calibri', bold=True, color='FFFFFF', size=16),
              fill=FILL_VARROC, alignment=ALIGN_CENTER)
    _row_height(ws, 1, 36)

    ws.merge_cells('A2:H2')
    _set_cell(ws, 2, 1, f"Drawing: {data.get('drawingId','—')}  |  "
              f"Scan: {data.get('scanDate','—')}  |  Engineer: {data.get('engineer','—')}",
              font=Font(name='Calibri', color='FFFFFF', size=10),
              fill=PatternFill("solid", fgColor="E87722"), alignment=ALIGN_CENTER)
    _row_height(ws, 2, 20)

    # ── Metadata block ──
    meta_rows = [
        ("Drawing ID",       data.get('drawingId','—')),
        ("Scan Date",        data.get('scanDate','—')),
        ("Engineer",         data.get('engineer','—')),
        ("Standards",        ", ".join(data.get('standardsChecked', []))),
        ("Sign-Off Status",  "❌ BLOCKED" if blocks else "✅ CLEARED"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=4):
        _set_cell(ws, i, 1, k, font=FONT_BOLD,   fill=FILL_SUBHEADER,
                  alignment=ALIGN_LEFT, border=THIN_BORDER)
        _set_cell(ws, i, 1, k, font=Font(name='Calibri', bold=True, color='FFFFFF', size=10),
                  fill=FILL_SUBHEADER, alignment=ALIGN_LEFT, border=THIN_BORDER)
        cell = ws.cell(row=i, column=2, value=v)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if k == "Sign-Off Status":
            cell.fill = FILL_FAIL if blocks else FILL_PASS
            cell.font = Font(name='Calibri', bold=True,
                             color='C0392B' if blocks else '27AE60', size=10)

    # ── Counts block ──
    ws.merge_cells('D4:E4')
    _set_cell(ws, 4, 4, "Violation Counts", font=FONT_WHITE_HDR, fill=FILL_HEADER,
              alignment=ALIGN_CENTER, border=THIN_BORDER)

    count_data = [
        ("Critical",  total_c,  "C0392B"),
        ("Major",     total_m,  "E67E22"),
        ("Minor",     total_n,  "D4AC0D"),
        ("Total",     total_c + total_m + total_n, "1B4F72"),
    ]
    for i, (label, val, col) in enumerate(count_data, start=5):
        _set_cell(ws, i, 4, label,
                  font=Font(name='Calibri', bold=True, color=col, size=11),
                  alignment=ALIGN_CENTER, border=THIN_BORDER)
        _set_cell(ws, i, 5, val,
                  font=Font(name='Calibri', bold=True, color=col, size=14),
                  alignment=ALIGN_CENTER, border=THIN_BORDER)

    # ── Column widths ──
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

    # ── Bar Chart ──
    chart_data = [["Severity", "Count"],
                  ["Critical", total_c],
                  ["Major",    total_m],
                  ["Minor",    total_n]]
    for ri, row in enumerate(chart_data, start=12):
        for ci, val in enumerate(row, start=4):
            ws.cell(row=ri, column=ci, value=val)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Violations by Severity"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    chart.style = 10
    chart.width  = 12
    chart.height = 8

    data_ref   = Reference(ws, min_col=5, min_row=12, max_row=15)
    cats_ref   = Reference(ws, min_col=4, min_row=13, max_row=15)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Colour each bar
    bar_colours = ['C0392B', 'E67E22', 'D4AC0D']
    for i, colour in enumerate(bar_colours):
        pt = DataPoint(idx=i)
        pt.spPr = None
        chart.series[0].dPt.append(pt)

    ws.add_chart(chart, "D17")

    # Footer note
    ws.merge_cells('A32:H32')
    _set_cell(ws, 32, 1,
              "Generated by VARROC AI Validation System | Not a substitute for qualified engineering review",
              font=Font(name='Calibri', italic=True, color='566573', size=8),
              alignment=ALIGN_CENTER)


# ── Generic issues sheet ───────────────────────────────────────────────────────
def build_issues_sheet(ws, violations, title, severity_label):
    ws.title = title
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(VIOLATION_COLS))}1')
    _set_cell(ws, 1, 1, f"{severity_label} Issues — {title}",
              font=Font(name='Calibri', bold=True, color='FFFFFF', size=13),
              fill=FILL_VARROC, alignment=ALIGN_CENTER)
    _row_height(ws, 1, 28)

    _write_violation_headers(ws, row=2)

    for i, v in enumerate(violations, start=3):
        _write_violation_row(ws, i, v, alt=(i % 2 == 0))

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(VIOLATION_COLS))}2"

    # Footer
    foot_row = len(violations) + 4
    ws.merge_cells(f'A{foot_row}:{get_column_letter(len(VIOLATION_COLS))}{foot_row}')
    _set_cell(ws, foot_row, 1,
              "Generated by VARROC AI Validation System | Not a substitute for qualified engineering review",
              font=Font(name='Calibri', italic=True, color='566573', size=8),
              alignment=ALIGN_CENTER)


# ── Main generator function ───────────────────────────────────────────────────
def generate_excel(data: dict, output_path: str = None) -> bytes:
    """
    Generate an Excel report from validation results.
    Args:
        data: dict matching /validate/full response schema
        output_path: optional file path to save.
    Returns:
        Excel bytes
    """
    wb = Workbook()

    all_violations = data.get('violations', [])
    critical = [v for v in all_violations if v.get('severity') == 'Critical']
    major    = [v for v in all_violations if v.get('severity') == 'Major']
    minor    = [v for v in all_violations if v.get('severity') == 'Minor']

    # Sheet 1: Summary
    ws_summary = wb.active
    build_summary_sheet(ws_summary, data)

    # Sheet 2: Critical
    ws_crit = wb.create_sheet("Critical Issues")
    build_issues_sheet(ws_crit, critical, "Critical Issues", "⛔ Critical")

    # Sheet 3: Major
    ws_major = wb.create_sheet("Major Issues")
    build_issues_sheet(ws_major, major, "Major Issues", "⚠ Major")

    # Sheet 4: All Issues
    ws_all = wb.create_sheet("All Issues")
    build_issues_sheet(ws_all, all_violations, "All Issues", "All")

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    if output_path:
        with open(output_path, 'wb') as f:
            f.write(excel_bytes)
        print(f"[excel_generator] Saved -> {output_path}")

    return excel_bytes


# ── CLI self-test ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    sample_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'sample_violations.json')
    with open(sample_path) as f:
        data = json.load(f)
    out = os.path.join(os.path.dirname(__file__), '..', '..', 'sample_report.xlsx')
    generate_excel(data, output_path=out)
    print(f"Sample Excel saved to: {os.path.abspath(out)}")
